"""Exportar os dados do portfólio para um excel para análise.

"""
from datetime import datetime
import numpy as np
import pandas as pd


def Func_criar_dfs_portifolio(df_exportar_f, indices_benchmark):
    """Creates DataFrames related to portfolio allocation and performance.

    Parameters:
        df_exportar_f (pd.DataFrame): DataFrame containing portfolio export data.
        indices_benchmark (list): Dates of portfolio rebalances.

    Returns:

        df_exportar2 (pd.DataFrame): DataFrame of portfolio allocation.
        df_exportar_pct (pd.DataFrame): DataFrame of percentage portfolio allocation.
        df_performance_ativo (pd.DataFrame): DataFrame of asset performance.
        df_diff_ativo (pd.DataFrame): DataFrame of allocation differences during the investment window.

    Notes:
        The function takes a DataFrame containing portfolio export data and a list of rebalance dates, then creates several DataFrames related to portfolio allocation and performance.

    Examples:
        >>> df_exportar2, df_exportar_pct, df_performance_ativo, df_diff_ativo = Func_criar_dfs_portifolio(df_exportar, indices_benchmark)
    """
    ## Criando o df de alocação da carteira
    df_exportar2 = df_exportar_f.copy()
    df_exportar2 = df_exportar2.dropna(axis=1, how="all")
    colunas = df_exportar2.columns
    colunas = colunas.sort_values()
    colunas = colunas.drop(["Montante", "LFTS3"])
    colunas = colunas.insert(0, "Montante")
    colunas = colunas.insert(1, "LFTS3")
    ## Reordenar colunas e selecionar as linhas de mudança de alocação
    df_exportar2 = df_exportar2.loc[indices_benchmark, colunas]

    ## Criando o df das diferenças de alocação durante a janela de investimento
    lista_colunas = df_exportar2.columns
    df_diff_ativo = pd.DataFrame()
    for coluna in lista_colunas:
        ativo = df_exportar2.loc[:, coluna]
        ativo_diff = ativo.diff()
        # Identificar as duplicatas no índice
        duplicates_mask = ativo_diff.index.duplicated(keep="first")
        # Criar uma nova série com base nas duplicatas
        new_series = ativo_diff[~duplicates_mask]
        new_series = new_series.shift(-1)
        new_series = new_series.dropna()
        df_diff_ativo[coluna] = new_series

    # Criando df da performance dos ativos
    df_performance_ativo = pd.DataFrame(columns=lista_colunas)

    # Rentabilidade do período total
    # Último dia do último ano
    ultm_data = df_diff_ativo.index.max()
    ano_ultm_data = ultm_data.year
    ultm_dia_ano = datetime(ano_ultm_data, 12, 31)

    # Criar um dataframe com a rentabilidade
    df_performance_ativo.loc[ultm_dia_ano, :] = df_diff_ativo.loc[:].sum(axis=0)

    # Ver a rentabilidade de todos os anos
    primeiro_dia = df_diff_ativo.index.min()
    primeiro_dia_ano = primeiro_dia.year

    # Iterando sobre os anos
    for ano in range(primeiro_dia_ano, ano_ultm_data + 1):
        df_temp = df_diff_ativo.loc[str(ano), :]
        df_performance_ativo.loc[datetime(ano, 12, 30), :] = df_temp.sum(axis=0)

    df_performance_ativo = df_performance_ativo.sort_index()

    # Alocação Percentual
    df_exportar_pct = df_exportar2.apply(
        lambda x: (x / df_exportar2["Montante"]) * 100,
        axis=0,
    )

    return df_exportar2, df_exportar_pct, df_performance_ativo, df_diff_ativo


def Func_search_alocation(
    row,
    df_exportar2,
    df_exportar_pct,
    df_diff_ativo,
    diff_df_benchmark,
):
    """Searches for allocation information and calculates returns for a given asset.

    Parameters:
        row (pd.Series): Row containing information about the asset.
        df_exportar2 (pd.DataFrame): DataFrame of portfolio allocation.
        df_exportar_pct (pd.DataFrame): DataFrame of percentage portfolio allocation.
        df_diff_ativo (pd.DataFrame): DataFrame of allocation differences during the investment window.
        diff_df_benchmark (pd.DataFrame): DataFrame of benchmark returns.

    Returns:
        pd.Series: Row with updated information about the asset including allocation and returns.

    Notes:
        The function takes a row of asset information and calculates various metrics, including allocation percentage in the portfolio, return percentage of the asset in relation to the portfolio, and return percentage of the asset.

    Examples:
        >>> updated_row = Func_search_alocation(row, df_exportar2, df_exportar_pct, df_diff_ativo, diff_df_benchmark)
    """
    ativo = row["Ticker"]
    data = row.name

    ## Selecionar a alocação do ativo dentro de df_exportar_pct
    condicao_data = df_exportar_pct.index == data
    try:
        Pct_ativo = df_exportar_pct.loc[condicao_data, ativo].values[-1]
        Pct_ativo = np.nan_to_num(Pct_ativo, nan=0)
        row["Aloc_Pct"] = Pct_ativo / 100
    except:
        Pct_ativo = 0

    # Calcular a rentabilidade percentual do ativo em relação a carteira

    try:
        delta_ativo = df_diff_ativo.loc[data, ativo]
        montante_carteira = df_exportar2.loc[data, "Montante"].values[-1]
    except AttributeError:
        delta_ativo = df_diff_ativo.loc[data, ativo]
        montante_carteira = df_exportar2.loc[data, "Montante"]
    except KeyError:
        delta_ativo = 0
        montante_carteira = 1
    finally:
        delta_ativo = np.nan_to_num(delta_ativo, nan=0)
        montante_carteira = np.nan_to_num(montante_carteira, nan=1)

        Ret_Pct_cart = delta_ativo / montante_carteira
        row["Ret_Pct_cart"] = Ret_Pct_cart

    # Calcular a rentabilidade percentual do ativo
    Ret_Pct_ativo = diff_df_benchmark.loc[data, ativo]
    Ret_Pct_ativo = np.nan_to_num(Ret_Pct_ativo, nan=0)
    row["Ret_Pct_ativo"] = Ret_Pct_ativo

    # Calcular a rentabilidade percentual do ativo
    return row


def Func_alterar_df_principal(
    info_verificao_total,
    df_benchmark,
    df_exportar,
    indices_benchmark,
):
    """Updates the main DataFrame with allocation and return information.

    Parameters:
        info_verificao_total (list): List containing DataFrames with asset information.
        df_benchmark (pd.DataFrame): DataFrame with benchmark data.
        df_exportar (pd.DataFrame): DataFrame with portfolio export data.
        indices_benchmark (list): List of dates for rebalancing.

    Returns:
        Tuple: Updated list of DataFrames, DataFrame of portfolio allocations, DataFrame of percentage portfolio allocations, DataFrame of asset performance.

    Notes:
        The function takes a list of DataFrames containing asset information and updates the main DataFrame with allocation and return metrics. It also returns DataFrames with portfolio allocations, percentage portfolio allocations, and asset performance.

    Example:
        updated_info, df_exportar2, df_exportar_pct, df_performance_ativo = Func_alterar_df_principal(info_verificao_total, df_benchmark, df_exportar, indices_benchmark)
    """
    df_principal = info_verificao_total[0].copy()
    colunas_antigas = df_principal.columns
    novas_colunas = ["Aloc_Pct", "Ret_Pct_cart", "Ret_Pct_ativo"]
    colunas_reordenadas = (
        list(colunas_antigas[:2]) + novas_colunas + list(colunas_antigas[2:])
    )

    df_principal.loc[:, ["Aloc_Pct", "Ret_Pct_cart", "Ret_Pct_ativo"]] = 0

    # Criar df variação percentual dos ativos
    diff_df_benchmark = df_benchmark.loc[indices_benchmark, :].pct_change().shift(-1)

    # Receber df de alocação
    (
        df_exportar2,
        df_exportar_pct,
        df_performance_ativo,
        df_diff_ativo,
    ) = Func_criar_dfs_portifolio(df_exportar, indices_benchmark)

    # Alterar o df principal
    df_principal = df_principal.apply(
        Func_search_alocation,
        axis=1,
        args=(df_exportar2, df_exportar_pct, df_diff_ativo, diff_df_benchmark),
    )

    df_principal = df_principal.apply(
        Func_search_alocation,
        axis=1,
        args=(df_exportar2, df_exportar_pct, df_diff_ativo, diff_df_benchmark),
    )
    df_principal.loc[:, ["Aloc_Pct", "Ret_Pct_cart", "Ret_Pct_ativo"]].fillna(
        0,
        inplace=True,
    )

    # Reordenando as colunas
    df_principal = df_principal[colunas_reordenadas]

    # Alterando o info_verificao_total[0]
    info_verificao_total[0] = df_principal.copy()
    return info_verificao_total, df_exportar2, df_exportar_pct, df_performance_ativo


def Func_editar_format_excel(excel_file):
    """Formats specified worksheets in an Excel file.

    Parameters:
        excel_file (str): Path to the Excel file.

    Returns:
        None: The function modifies the Excel file in-place.

    Notes:
        This function takes the path to an Excel file, loads the specified worksheets, and formats the cells with the desired styles, including font, fill, alignment, and borders. It also sets specific number formats for certain columns and adjusts column widths automatically.

    Examples:
        >>> Func_editar_format_excel('path/to/your/excel_file.xlsx')
    """
    # Importando as Bibliotecas
    import openpyxl as op
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers

    # Ler o arquivo excel
    wb = op.load_workbook(excel_file)

    # Ler os sheets
    ws_ret = wb["Retorno"]
    ws_aloc_dinheiro = wb["Rentabilidade_dinheiro"]

    # Lista de sheets
    lista_ws = [ws_ret]
    ## Iterar sobre os dataframes e planilhas
    for i in range(len(lista_ws)):
        # Selecionar o worksheet
        ws_excel = lista_ws[i]

        # Cabeçalho
        first_row = ws_excel[1]
        # Defina os estilos desejados para a primeira linha
        # Aplique os estilos às células da primeira linha
        for cell in first_row:
            cell.font = Font(name="Times New Roman", size=14, bold=True, color="000000")
            cell.fill = PatternFill(
                start_color="FFFFFFFF",
                end_color="FFFFFFFF",
                fill_type="solid",
            )
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000"),
            )

        # Formatação das células
        # Formatação da Segunda linha em diante

        for row in ws_excel.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(
                    name="Times New Roman",
                    size=12,
                    bold=False,
                    color="000000",
                )
                cell.fill = PatternFill(
                    start_color="FFFFFFFF",
                    end_color="FFFFFFFF",
                    fill_type="solid",
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"),
                )

        # Primeira coluna
        for row in ws_excel.iter_rows(min_row=2, min_col=1, max_col=1):
            for cell in row:
                cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2

        # Terceira coluna
        for row in ws_excel.iter_rows(min_row=2, min_col=3):
            for cell in row:
                cell.number_format = numbers.FORMAT_PERCENTAGE_00

        ## Iterar sobre as colunas e ajustar a largura automaticamente
        for column_cells in ws_excel.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws_excel.column_dimensions[column_cells[0].column_letter].width = (
                length + 3
            ) * 1.3

    # Lista de sheets parte 2
    lista_ws2 = [ws_aloc_dinheiro]
    ## Iterar sobre os dataframes e planilhas
    for i in range(len(lista_ws2)):
        # Selecionar o worksheet
        ws_excel = lista_ws2[i]

        # Cabeçalho
        first_row = ws_excel[1]
        # Defina os estilos desejados para a primeira linha
        # Aplique os estilos às células da primeira linha
        for cell in first_row:
            cell.font = Font(name="Times New Roman", size=14, bold=True, color="000000")
            cell.fill = PatternFill(
                start_color="FFFFFFFF",
                end_color="FFFFFFFF",
                fill_type="solid",
            )
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000"),
            )

        # Formatação das células
        # Formatação da Segunda linha em diante

        for row in ws_excel.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(
                    name="Times New Roman",
                    size=12,
                    bold=False,
                    color="000000",
                )
                cell.fill = PatternFill(
                    start_color="FFFFFFFF",
                    end_color="FFFFFFFF",
                    fill_type="solid",
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(border_style="thin", color="000000"),
                    right=Side(border_style="thin", color="000000"),
                    top=Side(border_style="thin", color="000000"),
                    bottom=Side(border_style="thin", color="000000"),
                )

        # Primeira coluna
        for row in ws_excel.iter_rows(min_row=2, min_col=1, max_col=1):
            for cell in row:
                cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2

        for row in ws_excel.iter_rows(min_row=2, min_col=2):
            for cell in row:
                cell.number_format = numbers.FORMAT_NUMBER_00

        ## Iterar sobre as colunas e ajustar a largura automaticamente
        for column_cells in ws_excel.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws_excel.column_dimensions[column_cells[0].column_letter].width = (
                length + 2
            ) * 1.1

    # Salvando e fechando a planilha
    wb.save(excel_file)
    wb.close()
    return None


def Func_exportar_excel(
    info_verificao_total,
    df_benchmark,
    df_exportar,
    indices_benchmark,
    excel_file="Alocação_seraaa.xlsx",
):
    """Exports data to an Excel file and formats the sheets.

    Parameters:
        info_verificao_total (list): List containing DataFrames with allocation information.
        df_benchmark (pd.DataFrame): DataFrame with benchmark data.
        df_exportar (pd.DataFrame): DataFrame with export data.
        indices_benchmark (list): List of dates for benchmark rebalancing.
        excel_file (str, optional): Path to the Excel file to be created. Defaults to "Alocação_seraaa.xlsx".

    Returns:
        None: The function creates and formats the Excel file in-place.

    Notes:
        This function exports specified DataFrames to an Excel file with multiple sheets. It uses the ExcelWriter to create the Excel file and then formats the sheets using the `Func_editar_format_excel` function.

    Example:
        Func_exportar_excel(info_verificao_total, df_benchmark, df_exportar, indices_benchmark, excel_file="Alocação_seraaa.xlsx")
    """
    (
        info_verificao_total,
        df_exportar2,
        df_exportar_pct,
        df_performance_ativo,
    ) = Func_alterar_df_principal(
        info_verificao_total,
        df_benchmark,
        df_exportar,
        indices_benchmark,
    )

    # Usando o ExcelWriter, cria um doc .xlsx, usando engine='xlsxwriter'
    writer = pd.ExcelWriter(excel_file, engine="xlsxwriter")

    # Aba de Retorno
    info_verificao_total[0].to_excel(writer, sheet_name="Retorno", index=True)

    # Aba de Performance
    df_performance_ativo.to_excel(
        writer,
        sheet_name="Rentabilidade_dinheiro",
        index=True,
    )

    # Fecha o ExcelWriter e gera o arquivo .xlsx
    writer.close()

    # Editar o arquivo excel
    Func_editar_format_excel(excel_file)
    return None
